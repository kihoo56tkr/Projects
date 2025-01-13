import openpyxl
from datetime import datetime, timedelta
import statistics

## File paths: global variables
reportTemplate = r'\\.xlsx'
database = r'\\.xlsx'

## For Testing
#today = "2024-10-09 09:00:37"
#formatdt = '%Y-%m-%d %H:%M:%S'
#today = datetime.strptime(today,formatdt)
today = datetime.today()
today_day = today.strftime("%d")
today_month = today.strftime("%B")
today_year = int(today.strftime("%Y"))
formatted_date = today.strftime("%d-%m-%y")
 
prev = today - timedelta(days = 20) # Assumes code will run within first 2 weeks of new quarter
prev_day = prev.strftime("%d")
prev_month = prev.strftime("%B")
prev_month_num = int(prev.strftime("%m"))
prev_year = int(prev.strftime("%Y"))
 
Quarter = dict(January=4, February=4, March=4, April=1, May=1, June=1, July=2, August=2, September=2, October=3, November=3, December=3)

if Quarter[prev_month] == 4:
    refsheet = str(Quarter[prev_month]) + "Q" + str(prev_year - 1)[2:]
else:
    refsheet = str(Quarter[prev_month]) + "Q" + str(prev_year)[2:]
 if Quarter[today_month] == 4:
     newsheet = str(Quarter[today_month]) + "Q" + str(today_year - 1)[2:]
     emailsub = str(Quarter[today_month]) + "Qtr" + str(today_year - 1)
else:
    newsheet = str(Quarter[today_month]) + "Q" + str(today_year)[2:]
    emailsub = str(Quarter[today_month]) + "Qtr" + str(today_year)

print("New Sheet: " + newsheet)

print("Sheet to reference: " + refsheet)
 
def get_last_date_of_quarter(year, month):
    if month in [1, 2, 3]:
        return datetime(year, 3, 31)
    elif month in [4, 5, 6]:
        return datetime(year, 6, 30)
    elif month in [7, 8, 9]:
        return datetime(year, 9, 30)
    elif month in [10, 11, 12]:
        return datetime(year, 12, 31)
 
def duplicate_and_modify_sheet(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Check if previous sheet exists
    if refsheet in workbook.sheetnames:
        # Check if new sheet already exists and remove it
        if newsheet in workbook.sheetnames:
            std_workbook = workbook[newsheet]
            workbook.remove(std_workbook)
            print("Existing " + newsheet + " sheet found and removed.")
        # Reference the previous sheet
        sheet_ref = workbook[refsheet]
        # Duplicate the previous sheet
        sheet_new = workbook.copy_worksheet(sheet_ref)
        sheet_new.title = newsheet  # Rename the duplicated sheet
        
        # Set cell values in New Sheet from Old Sheet
        sheet_new['C5'] = sheet_ref['O5'].value
        print(sheet_ref['O5'].value)
        sheet_new['C7'] = sheet_ref['O7'].value
        sheet_new['C2'] = "As at start of Q" + str(Quarter[today_month]) + " " + str(newsheet)[2:]
        sheet_new['O2'] = "As at end of Q" + str(Quarter[today_month]) + " " + str(newsheet)[2:]
        sheet_new['D1'] = "Applications received or processed in Q" + str(Quarter[today_month]) + " " + str(emailsub[-4:]) + " [2], [3]"
        sheet_new['A10'] = "[2] Applications processed in Q" + str(Quarter[today_month]) + " " + str(emailsub[-4:]) + " can include applications received in previous quarters. "

        # Save the workbook with the new sheet
        workbook.save(file_path)
        print("Sheet " + refsheet + " was duplicated to " + newsheet + " and modified successfully.")
    else:
        print("Sheet " + refsheet + " not found in the workbook.")

def count_entries(file_path):
    # Load the workbook and sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Assumes data is in the active sheet
    count_bank = 0
    count_ro = 0

    for row in sheet.iter_rows(min_row=4):  # Start from row 4
        status = row[3].value  # Column D
        date_val = row[1].value  # Column B
        #print(date_val)
        # Ensure that the date_val is a datetime object
        if isinstance(date_val, datetime):
            month = date_val.month
            year = int(date_val.year)
            #print(prev_month_num - 2, prev_month_num, prev_year)

            if prev_month_num - 2 <= month <= prev_month_num and year == prev_year:
                #print(month, year)
                if status != "RO":
                    count_bank += 1
                elif status == "RO":
                    count_ro += 1

    print("Done count_entries")
    return count_bank, count_ro

def update_counts(source_file, target_file, bank_count, ro_count):
    # Load the target workbook and sheet
    target_workbook = openpyxl.load_workbook(target_file)
    sheet_new = target_workbook[newsheet]
    # Update counts in specific cells
    sheet_new['D5'] = bank_count
    sheet_new['D7'] = ro_count
    # Save the target workbook
    target_workbook.save(target_file)

    print("Done update_counts")
 
def calculating_outstanding(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Check if Previous sheet exists
    if newsheet in workbook.sheetnames:
        # Reference the Previous sheet
        sheet_new = workbook[newsheet]
        # Set cell values in New sheet from Previous Sheet
        # print(sheet_new['O5'].value, sheet_new['C5'].value, sheet_new['D5'].value, sheet_new['E5'].value)
        sheet_new['O5'] = sheet_new['C5'].value + sheet_new['D5'].value - sheet_new['E5'].value
        sheet_new['O7'] = sheet_new['C7'].value + sheet_new['D7'].value - sheet_new['E7'].value
        # Save the workbook with the new sheet

        workbook.save(file_path)

        print("Done calculating_outstanding")

def update_processed(source_file, target_file):
    # Load the workbook and sheet
    processed_bank = 0
    processed_ro = 0
    processed = []

    workbook = openpyxl.load_workbook(source_file)
    sheet = workbook.active  # Assumes data is in the active sheet

    for row in sheet.iter_rows(min_row=4):  # Start from row 4
        month_processed = row[14].value  # Column O
        year_processed = row[15].value  # Column P
        status = row[11].value  # Column L
        status1 = row[3].value  # Column D
        cell = row[0].coordinate

        #print(month_processed, year_processed, status, status1, cell)

        if month_processed == None or year_processed == None or month_processed == "N.A.":
            continue
        try:
            int(month_processed)
        except:
            continue

        if prev_month_num - 2 <= month_processed <= prev_month_num and year_processed == prev_year:
            if status != "In Process":
                if status1 != "RO":
                    processed_bank += 1
                    processed.append(cell)
                elif status1 == "RO":
                    processed_ro += 1
                    processed.append(cell)

    # Load the target workbook and sheet
    target_workbook = openpyxl.load_workbook(target_file)
    sheet_new = target_workbook[newsheet]

    # Update counts in specific cells
    sheet_new['E5'] = processed_bank
    sheet_new['E7'] = processed_ro

    # Save the target workbook
    target_workbook.save(target_file)
    print("Done update_processed")

    return processed

def update_Met_Service_Standards(source_file, target_file, processed_cells):    
    target_workbook = openpyxl.load_workbook(target_file)
    new_sheet = target_workbook[newsheet]  
    workbook = openpyxl.load_workbook(source_file, data_only=True)

    sheet = workbook.active  # Assumes data is in the active sheet

    all_data_ro = []
    all_data_wb = []

    time_taken_ro = []
    time_taken_wb = []

    number_mss_ro = []
    number_mss_wb = []
 
    for row in cells:
        start = row[1:]
        status = str(sheet["D" + str(start)].value)  # Column D
        time_taken = int(sheet["Q" + str(start)].value)

        #print(time_taken, status)
        if time_taken <= 122 and status != "RO":
            all_data_wb.append(100)
            time_taken_wb.append(time_taken)
            number_mss_wb.append(1)

        if time_taken > 122 and status != "RO":
            all_data_wb.append(0)  
            time_taken_wb.append(time_taken)
            number_mss_wb.append(0)

        if time_taken <= 61 and status == "RO":
            all_data_ro.append(100) 
            time_taken_ro.append(time_taken)
            number_mss_ro.append(1)

        if time_taken > 61 and status == "RO":
            all_data_ro.append(0) 
            time_taken_ro.append(time_taken)
            number_mss_wb.append(0)

    #print(all_data_ro, all_data_wb, number_mss_ro, number_mss_wb)

    if all_data_ro == []:
        percentage_ro = "-"        
    elif 100 not in all_data_ro:
        percentage_ro = "0.0%"
    else:
        percentage_ro = round(sum(all_data_ro)/len(all_data_ro), 1)
        percentage_ro = str(percentage_ro) + "%"

    if all_data_wb == []:
        percentage_wb = "-"         
    elif 100 not in all_data_wb:
        percentage_wb = "0.0%"
    else:
        percentage_wb = round(sum(all_data_wb)/len(all_data_wb), 1)
        percentage_wb = str(percentage_wb) + "%"

    #print(percentage_ro, percentage_wb)
 
    new_sheet['G5'] = percentage_wb
    new_sheet['G7'] = percentage_ro

    target_workbook.save(target_file)

    print("Done update_Met_Service_Standards")
 
    return time_taken_ro, time_taken_wb, number_mss_ro, number_mss_wb
 
def update_Number_Met_Service_Standards(target_file, number_ro, number_wb):
    #print(number_ro, number_wb)
    target_workbook = openpyxl.load_workbook(target_file)
    new_sheet = target_workbook[newsheet] 

    if number_wb == []:
        new_sheet['F5'] = "-"
    else:
        new_sheet['F5'] = sum(number_wb)       
    if number_ro == []:
        new_sheet['F7'] = "-"
    else:
        new_sheet['F7'] = sum(number_ro)  
    target_workbook.save(target_file)

    print("Done update_Number_Met_Service_Standards")

def Update_Mean_n_Median(target_file, time_taken_ro, time_taken_wb):
    target_workbook = openpyxl.load_workbook(target_file)
    new_sheet = target_workbook[newsheet]  
    if time_taken_ro == []:
        mean_ro, median_ro = "-", "-"
        #print(None)
    else:
        time_taken_ro_30 = [x / 30 for x in time_taken_ro] # each month is 30 days
        mean_ro = round(sum(time_taken_ro_30)/len(time_taken_ro_30), 1)
        median_ro = round(statistics.median(time_taken_ro_30), 1) 
        #print(time_taken_ro_30)

    if time_taken_wb == []:
        mean_wb, median_wb = "-", "-"
        #print(None)
    else:
        time_taken_wb_30 = [x / 30 for x in time_taken_wb] # each month is 30 days
        mean_wb = round(sum(time_taken_wb_30)/len(time_taken_wb_30), 1)
        median_wb = round(statistics.median(time_taken_wb_30), 1) 
        #print(time_taken_wb_30)

    new_sheet['H5'] = mean_wb
    new_sheet['H7'] = mean_ro
    new_sheet['I5'] = median_wb
    new_sheet['I7'] = median_ro

    target_workbook.save(target_file)

    print("Done Update_Mean_n_Median")

def get_past_due(source_file, target_file, prev_month_num, prev_year):
    # Load the workbook and sheet
    workbook = openpyxl.load_workbook(source_file, data_only=True)  # Set here to be True to get the data_only
    sheet = workbook.active  # Assumes data is in the active sheet

    processed_bank = 0
    processed_ro = 0
    count_bank_outstanding = 0
    count_ro_outstanding = 0
 
    last_date_of_quarter = get_last_date_of_quarter(prev_year, prev_month_num)
 
    for row in sheet.iter_rows(min_row=4):  # Start from row 4
        status = row[11].value  # Column L
        status1 = row[3].value  # Column D
        date_val = row[12].value  # Column M
        # Ensure that the date_val is a datetime object
        if isinstance(date_val, datetime):
            # Compare the last date of the quarter with the year with date_val dat, month and year
            ## rewrite the code below
            #print(last_date_of_quarter)
            #print(date_val)

            days_diff = (last_date_of_quarter - date_val).days

            if status == "In Process":
                #print(days_diff)
                if status1 != "RO":  # if it is a bank
                    processed_bank += 1
                    if days_diff > 122:
                        count_bank_outstanding += 1
                elif status1 == "RO":  # if it is a Rep office
                    processed_ro += 1
                    if days_diff > 61:
                        count_ro_outstanding += 1
 
    # Load the target workbook and sheet
    target_workbook = openpyxl.load_workbook(target_file)
    sheet_new = target_workbook[newsheet]
    # Update counts in specific cells
    bank_percentage = (count_bank_outstanding / sheet_new['O5'].value) * 100 if sheet_new['O5'].value > 0 else "-"
    ro_percentage = (count_ro_outstanding / sheet_new['O7'].value) * 100 if sheet_new['O7'].value > 0 else "-"
 
    try:
        int(bank_percentage)
    except:
        sheet_new['P5'] = bank_percentage
    else:
        sheet_new['P5'] = f"{bank_percentage:.1f}%"

    try:
        int(ro_percentage)
    except:
        sheet_new['P7'] = ro_percentage
    else:
        sheet_new['P7'] = f"{ro_percentage:.1f}%"

    sheet_new['Q5'] = count_bank_outstanding
    sheet_new['Q7'] = count_ro_outstanding
 
    # Save the target workbook
    target_workbook.save(target_file)

    print("Done get_past_due")
 
def update_Annual_Mean_n_Median(source_file, target_file):
    # Load the workbook and sheet
    processed_bank = []
    processed_ro = []
    workbook = openpyxl.load_workbook(source_file, data_only=True)

    sheet = workbook.active  # Assumes data is in the active sheet

    for row in sheet.iter_rows(min_row=4):  # Start from row 4
        year_processed = row[15].value  # Column P
        status = row[11].value  # Column L
        status1 = row[3].value  # Column D
        time_taken = row[16].value # Column Q

        #print(time_taken, year_processed, status, status1)

        try:
            int(year_processed)
        except:
            continue

        if int(year_processed) == prev_year:
            if status != "In Process":
                if status1 != "RO":
                    processed_bank.append(int(time_taken))
                elif status1 == "RO":
                    processed_ro.append(int(time_taken))

    # Load the target workbook and sheet

    target_workbook = openpyxl.load_workbook(target_file)
    new_sheet = target_workbook[newsheet]  

    if processed_ro == []:
        mean_ro, median_ro = "-", "-"
        #print(None)
    else:
        time_taken_ro_30 = [x / 30 for x in processed_ro] # each month is 30 days
        mean_ro = round(sum(time_taken_ro_30)/len(time_taken_ro_30), 1)
        median_ro = round(statistics.median(time_taken_ro_30), 1) 
        print(time_taken_ro_30)

    if processed_bank == []:
        mean_wb, median_wb = "-", "-"
        #print(None)
    else:
        time_taken_wb_30 = [x / 30 for x in processed_bank] # each month is 30 days
        mean_wb = round(sum(time_taken_wb_30)/len(time_taken_wb_30), 1)
        median_wb = round(statistics.median(time_taken_wb_30), 1) 
        print(time_taken_wb_30)

    new_sheet['J5'] = mean_wb
    new_sheet['J7'] = mean_ro
    new_sheet['K5'] = median_wb
    new_sheet['K7'] = median_ro

    target_workbook.save(target_file)

    print("Done update_Annual_Mean_n_Median")

def temp(target_file): # Fills up other cells with "-"
    # Load the workbook and sheet
    target_workbook = openpyxl.load_workbook(target_file)
    sheet_new = target_workbook[newsheet]

    sheet_new['J5'] = "-"
    sheet_new['J7'] = "-"   
    sheet_new['K5'] = "-"
    sheet_new['K7'] = "-"   
    sheet_new['L5'] = "-"
    sheet_new['L7'] = "-"  
    sheet_new['M5'] = "-"
    sheet_new['M7'] = "-"  
    sheet_new['N5'] = "-"
    sheet_new['N7'] = "-"  

    target_workbook.save(target_file)

    print("Done temp")

# Provide the path to your Excel file
duplicate_and_modify_sheet(reportTemplate)
# Get counts from source file
bank_count, ro_count = count_entries(database)
# Update counts in target file
cells = update_processed(database, reportTemplate)
#print(cells)
update_counts(database, reportTemplate, bank_count, ro_count)
calculating_outstanding(reportTemplate)
time_taken_ro, time_taken_wb, number_mss_ro, number_mss_wb = update_Met_Service_Standards(database, reportTemplate, cells)
update_Number_Met_Service_Standards(reportTemplate, number_mss_ro, number_mss_wb)
Update_Mean_n_Median(reportTemplate, time_taken_ro, time_taken_wb)
get_past_due(database, reportTemplate, prev_month_num, prev_year)
temp(reportTemplate)
if Quarter[today_month] == 4:
    update_Annual_Mean_n_Median(database, reportTemplate)
else:
    print("Not Q4 so no Annual Mean n Median")
print("DONE UPDATING")  
 